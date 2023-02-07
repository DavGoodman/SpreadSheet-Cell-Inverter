#! pyhton3
# SpreadSheetCellInverter.py inverts columns and rows in an excel spreadsheet

import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
wb = openpyxl.load_workbook("write_formula.xlsx", data_only=True)
sheet = wb['Sheet']

sheet_data = []
for y in range(1, sheet.max_column + 1):
    column = []
    for x in range(1, sheet.max_row + 1):
        cell = sheet[get_column_letter(y) + str(x)].value
        column.append(cell)

    sheet_data.append(column)

wb.close()

wb1 = openpyxl.Workbook()
sheet = wb1["Sheet"]
print(sheet_data)


for x, row in enumerate(sheet_data):
    for y, column in enumerate(sheet_data[x]):
        sheet[get_column_letter(y + 1) + str(x + 1)] = column

wb1.save("inverted.xlsx")

